from typing import override
from django.contrib.auth.forms import UserCreationForm
from django.views import View
from django.core.mail import send_mail
from django.conf import settings
from django.views.generic import ListView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.contrib.auth import logout
from django.contrib import messages
from .forms import CustomUserCreationForm, CourseForm, CalificacionesForm, NotaIndividualFormSet # Importa NotaIndividualFormSet
from django.contrib.auth.models import Group, User
from django.db.models import Q
from datetime import date, datetime # Añade 'datetime' aquí
from django.db.models import Q, Max # Asegúrate de que 'Max' esté importado
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from accounts.models import Registro

# Create your views here.

class UsuariosRegistradosView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Registro
    template_name = 'usuarios_registrados.html'
    context_object_name = 'registros'

    def test_func(self):
        return self.request.user.groups.filter(name='Secretaría').exists()

    def get_queryset(self):
        queryset = Registro.objects.all().select_related('user').prefetch_related('user__groups')
        # Eliminar o comentar las siguientes líneas si ya no necesitas el filtro por grupo desde la URL
        # grupo = self.request.GET.get('grupo')
        search = self.request.GET.get('search')

        # Asegurarse de que solo se muestren los estudiantes
        queryset = queryset.filter(user__groups__name='Estudiantes')

        # if grupo:
        #     queryset = queryset.filter(user__groups__name=grupo)

        if search:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search)
            )

        return queryset


@login_required
def export_usuarios_excel(request):
    search_query = request.GET.get('search', '')
    registros = Registro.objects.filter(user__groups__name='Estudiantes')

    if search_query:
        registros = registros.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(carnet__icontains=search_query)
        ).distinct()

    context = {
        'registros': registros
    }
    excel_file = generate_excel(context)
    response = HttpResponse(excel_file.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="usuarios_registrados.xlsx"'
    return response

@login_required
def export_matriculas_pdf(request):
    curso_academico_id = request.GET.get('curso_academico')
    curso_id = request.GET.get('curso')
    student_id = request.GET.get('student')

    matriculas = Matriculas.objects.all()

    if curso_academico_id:
        matriculas = matriculas.filter(course__curso_academico__id=curso_academico_id)
    if curso_id:
        matriculas = matriculas.filter(course__id=curso_id)
    if student_id:
        matriculas = matriculas.filter(student__id=student_id)

    context = {
        'matriculas': matriculas,
        'curso_academico': CursoAcademico.objects.get(id=curso_academico_id) if curso_academico_id else None,
    }
    return render_to_pdf('matriculas_pdf.html', context)

@login_required
def export_matriculas_excel(request):
    curso_academico_id = request.GET.get('curso_academico')
    curso_id = request.GET.get('curso')
    student_id = request.GET.get('student')

    matriculas = Matriculas.objects.all()

    if curso_academico_id:
        matriculas = matriculas.filter(course__curso_academico__id=curso_academico_id)
    if curso_id:
        matriculas = matriculas.filter(course__id=curso_id)
    if student_id:
        matriculas = matriculas.filter(student__id=student_id)

    context = {
        'matriculas': matriculas,
        'curso_academico': CursoAcademico.objects.get(id=curso_academico_id) if curso_academico_id else None,
    }
    
    # Generar el archivo Excel
    output = generate_excel(context)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=matriculas.xlsx'
    return response

from django.views.generic import DetailView
from .models import CursoAcademico, Curso, Matriculas, Calificaciones, Asistencia

# Función auxiliar para generar PDF
def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None

# Función auxiliar para generar Excel
def generate_excel(context_dict={}):
    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    
    # Estilos para el Excel
    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Obtener datos del contexto
    curso_academico = context_dict.get('curso_academico')
    cursos = context_dict.get('cursos', [])
    matriculas = context_dict.get('matriculas', [])
    calificaciones = context_dict.get('calificaciones', [])
    asistencias = context_dict.get('asistencias', [])
    registros = context_dict.get('registros', []) # Añadir registros al contexto
    
    # Hoja de información general
    if curso_academico:
        ws_info = wb.active
        ws_info.title = "Información General"
        ws_info['A1'] = f"Curso Académico: {curso_academico.nombre}"
        ws_info['A2'] = f"Activo: {'Sí' if curso_academico.activo else 'No'}"
        ws_info['A3'] = f"Archivado: {'Sí' if curso_academico.archivado else 'No'}"
        ws_info['A4'] = f"Fecha de Creación: {curso_academico.fecha_creacion}"
    else:
        # Remove the default active sheet if no curso_academico is provided
        if 'Sheet' in wb.sheetnames:
            std = wb['Sheet']
            wb.remove(std)
    
    # Hoja de cursos
    if cursos:
        ws_cursos = wb.create_sheet(title="Cursos")
        # Encabezados
        headers = ["Nombre del Curso", "Profesor", "Estado"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_cursos.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos
        for row_num, curso in enumerate(cursos, 2):
            ws_cursos.cell(row=row_num, column=1, value=curso.name)
            ws_cursos.cell(row=row_num, column=2, value=curso.teacher.get_full_name() or curso.teacher.username)
            ws_cursos.cell(row=row_num, column=3, value=curso.get_status_display())
            
            # Aplicar bordes a todas las celdas
            for col_num in range(1, 4):
                ws_cursos.cell(row=row_num, column=col_num).border = border
        
        # Ajustar ancho de columnas
        for col in ws_cursos.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws_cursos.column_dimensions[column].width = adjusted_width
    
    # Hoja de matrículas
    if matriculas:
        ws_matriculas = wb.create_sheet(title="Matrículas")
        # Encabezados
        headers = ["Estudiante", "Curso Académico", "Curso", "Fecha Matrícula", "Estado Matrícula"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_matriculas.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos
        for row_num, matricula in enumerate(matriculas, 2):
            ws_matriculas.cell(row=row_num, column=1, value=matricula.student.get_full_name() or matricula.student.username)
            ws_matriculas.cell(row=row_num, column=2, value=matricula.course.curso_academico.nombre if matricula.course and matricula.course.curso_academico else 'N/A')
            ws_matriculas.cell(row=row_num, column=3, value=matricula.course.name if matricula.course else 'N/A')
            ws_matriculas.cell(row=row_num, column=4, value=matricula.fecha_matricula.strftime('%d/%m/%Y') if matricula.fecha_matricula else 'N/A')
            ws_matriculas.cell(row=row_num, column=5, value=matricula.get_estado_display())
            
            # Aplicar bordes a todas las celdas
            for col_num in range(1, 6):
                ws_matriculas.cell(row=row_num, column=col_num).border = border
        
        # Ajustar ancho de columnas
        for col in ws_matriculas.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws_matriculas.column_dimensions[column].width = adjusted_width

    
    # Hoja de calificaciones
    if calificaciones:
        ws_calificaciones = wb.create_sheet(title="Calificaciones")
        
        # Determinar el número máximo de notas individuales
        max_notas = 0
        for calificacion in calificaciones:
            num_notas = calificacion.notas.count()
            if num_notas > max_notas:
                max_notas = num_notas
        
        # Crear encabezados dinámicos
        headers = ["Estudiante", "Curso"]
        for i in range(1, max_notas + 1):
            headers.append(f"Nota {i}")
        headers.append("Promedio")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_calificaciones.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos
        for row_num, calificacion in enumerate(calificaciones, 2):
            # Información básica
            ws_calificaciones.cell(row=row_num, column=1, value=calificacion.student.get_full_name() or calificacion.student.username)
            ws_calificaciones.cell(row=row_num, column=2, value=calificacion.course.name)
            
            # Notas individuales
            notas_individuales = list(calificacion.notas.all().order_by('fecha_creacion'))
            for i, nota in enumerate(notas_individuales, 3):
                ws_calificaciones.cell(row=row_num, column=i, value=nota.valor)
            
            # Rellenar con N/A las notas que no existen
            for i in range(len(notas_individuales) + 3, max_notas + 3):
                ws_calificaciones.cell(row=row_num, column=i, value='N/A')
            
            # Promedio
            ws_calificaciones.cell(row=row_num, column=max_notas + 3, value=calificacion.average if calificacion.average is not None else 'N/A')
            
            # Aplicar bordes a todas las celdas
            for col_num in range(1, max_notas + 4):
                ws_calificaciones.cell(row=row_num, column=col_num).border = border
        
    # Hoja de usuarios registrados
    if registros:
        ws_usuarios = wb.create_sheet(title="Usuarios Registrados")
        # Encabezados
        headers = [
            "Nombre", "Apellidos", "Email", "Nacionalidad", "Carnet ID", "Sexo",
            "Dirección", "Municipio", "Provincia", "Movil", "Grado Académico",
            "Ocupación", "Título", "Grupo", "Fecha de Registro"
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws_usuarios.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos
        for row_num, registro in enumerate(registros, 2):
            ws_usuarios.cell(row=row_num, column=1, value=registro.user.first_name)
            ws_usuarios.cell(row=row_num, column=2, value=registro.user.last_name)
            ws_usuarios.cell(row=row_num, column=3, value=registro.user.email)
            ws_usuarios.cell(row=row_num, column=4, value=registro.nacionalidad)
            ws_usuarios.cell(row=row_num, column=5, value=registro.carnet)
            ws_usuarios.cell(row=row_num, column=6, value=registro.sexo)
            ws_usuarios.cell(row=row_num, column=7, value=registro.address)
            ws_usuarios.cell(row=row_num, column=8, value=registro.location)
            ws_usuarios.cell(row=row_num, column=9, value=registro.provincia)
            ws_usuarios.cell(row=row_num, column=10, value=registro.movil)
            ws_usuarios.cell(row=row_num, column=11, value=registro.get_grado_display())
            ws_usuarios.cell(row=row_num, column=12, value=registro.get_ocupacion_display())
            ws_usuarios.cell(row=row_num, column=13, value=registro.titulo)
            ws_usuarios.cell(row=row_num, column=14, value=registro.user.groups.first().name if registro.user.groups.first() else '')
            ws_usuarios.cell(row=row_num, column=15, value=registro.user.date_joined.strftime("%d/%m/%Y"))
            
            for col_num in range(1, len(headers) + 1):
                ws_usuarios.cell(row=row_num, column=col_num).border = border
        
        # Ajustar ancho de columnas
        for col in ws_usuarios.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws_usuarios.column_dimensions[column].width = adjusted_width

    if calificaciones:
        # Ajustar ancho de columnas
        for col in ws_calificaciones.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws_calificaciones.column_dimensions[column].width = adjusted_width

    # Hoja de asistencias
    if asistencias:
        ws_asistencias = wb.create_sheet(title="Asistencias")
        # Encabezados
        headers = ["Estudiante", "Curso", "Fecha", "Presente"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_asistencias.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Datos
        for row_num, asistencia in enumerate(asistencias, 2):
            ws_asistencias.cell(row=row_num, column=1, value=asistencia.student.get_full_name() or asistencia.student.username)
            ws_asistencias.cell(row=row_num, column=2, value=asistencia.course.name)
            ws_asistencias.cell(row=row_num, column=3, value=asistencia.date.strftime('%d/%m/%Y') if asistencia.date else 'N/A')
            ws_asistencias.cell(row=row_num, column=4, value='Sí' if asistencia.presente else 'No')

            # Aplicar bordes a todas las celdas
            for col_num in range(1, 5):
                ws_asistencias.cell(row=row_num, column=col_num).border = border

        # Ajustar ancho de columnas
        for col in ws_asistencias.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws_asistencias.column_dimensions[column].width = adjusted_width

    # Guardar el archivo en memoria
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file

class CursoAcademicoDetailView(DetailView):
    model = CursoAcademico
    template_name = 'curso_academico_detail.html'
    context_object_name = 'curso_academico'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        curso_academico = self.get_object()
        
        # Obtener los filtros de la URL
        curso_id = self.request.GET.get('curso')
        estudiante_id = self.request.GET.get('estudiante')
        
        # Filtrar cursos
        cursos = Curso.objects.filter(matriculas__curso_academico=curso_academico).distinct()
        if curso_id:
            cursos = cursos.filter(id=curso_id)
        context['cursos'] = cursos
        
        # Filtrar matrículas
        matriculas = Matriculas.objects.filter(curso_academico=curso_academico)
        if curso_id:
            matriculas = matriculas.filter(course_id=curso_id)
        if estudiante_id:
            matriculas = matriculas.filter(student_id=estudiante_id)
        context['matriculas'] = matriculas
        
        # Filtrar calificaciones
        calificaciones = Calificaciones.objects.filter(curso_academico=curso_academico)
        if curso_id:
            calificaciones = calificaciones.filter(course_id=curso_id)
        if estudiante_id:
            calificaciones = calificaciones.filter(student_id=estudiante_id)
        context['calificaciones'] = calificaciones
        
        # Filtrar asistencias
        asistencias = Asistencia.objects.filter(course__matriculas__curso_academico=curso_academico).distinct()
        if curso_id:
            asistencias = asistencias.filter(course_id=curso_id)
        if estudiante_id:
            asistencias = asistencias.filter(student_id=estudiante_id)
        context['asistencias'] = asistencias
        
        # Agregar listas para los selectores de filtro
        context['cursos_disponibles'] = Curso.objects.filter(matriculas__curso_academico=curso_academico).distinct()
        context['estudiantes_disponibles'] = User.objects.filter(
            matriculas__curso_academico=curso_academico
        ).distinct()
        
        return context
        
    def render_to_response(self, context, **response_kwargs):
        # Verificar si se solicita PDF
        if 'pdf' in self.request.GET:
            # Añadir fecha actual al contexto para el PDF
            context['now'] = datetime.now()
            # Crear un template para PDF basado en el mismo contexto
            pdf = render_to_pdf('curso_academico_pdf.html', context)
            if pdf:
                response = HttpResponse(pdf, content_type='application/pdf')
                filename = f"curso_academico_{context['curso_academico'].nombre}.pdf"
                content = f"attachment; filename={filename}"
                response['Content-Disposition'] = content
                return response
        # Verificar si se solicita Excel
        elif 'excel' in self.request.GET:
            # Generar archivo Excel
            excel_file = generate_excel(context)
            if excel_file:
                response = HttpResponse(excel_file.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                filename = f"curso_academico_{context['curso_academico'].nombre}.xlsx"
                response['Content-Disposition'] = f'attachment; filename={filename}'
                return response
        # Si no se solicita PDF ni Excel, renderizar normalmente
        return super().render_to_response(context, **response_kwargs)
from django.views.generic import TemplateView, CreateView, ListView, View
from .models import Calificaciones, Curso, Matriculas, CursoAcademico
from django.views.generic.edit import UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin # Importar LoginRequiredMixin


class BaseContextMixin:
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        group_name = None
        if user.is_authenticated:
            group = Group.objects.filter(user=user).first()
            if group:
                group_name = group.name
            context['group_name'] = group_name
        return context


class HomeView(BaseContextMixin, TemplateView):
    template_name = 'home.html'

    @override
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
        if curso_academico_activo:
            courses = Curso.objects.filter(curso_academico=curso_academico_activo)
        else:
            courses = Curso.objects.none()
        # Group courses into chunks of four for the carousel
        grouped_courses = [courses[i:i + 4] for i in range(0, len(courses), 4)]
        context['grouped_courses'] = grouped_courses
        student = self.request.user if self.request.user.is_authenticated else None

        for item in courses:
            if student:
                registration = Matriculas.objects.filter(
                    course=item, student=student).first()
                item.is_enrolled = registration is not None
            else:
                item.is_enrolled = False

            enrollment_count = Matriculas.objects.filter(course=item).count()
            item.enrollment_count = enrollment_count

        context['courses'] = courses
        return context



class ListadoCursosView(BaseContextMixin, ListView):
    model = Curso
    template_name = 'cursos.html'
    context_object_name = 'courses'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_queryset(self):
        # Obtener el CursoAcademico activo
        curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
        if curso_academico_activo:
            # Filtrar los cursos por el CursoAcademico activo
            return Curso.objects.filter(curso_academico=curso_academico_activo)
        return Curso.objects.none() # No mostrar cursos si no hay un CursoAcademico activo

# para cerrar sesion


def logout_view(request):
    logout(request)
    return redirect('principal:home')


# pagina de Registro
""" class RegisterView(View):
    def get(self, request):
        data = {
            'form': RegisterForm()
        }
        return render(request, 'registration/registro.html', data)

    def post(self, request):
        user_creation_form=RegisterForm(data=request.POST)
        if user_creation_form.is_valid():
            user_creation_form.save()  
            user = authenticate(username=user_creation_form.cleaned_data['username'],
                                password=user_creation_form.cleaned_data['password'])
            login(request, user)
            return redirect('principal:home')
        data = {
            'form':user_creation_form
        }   
        return render(request, 'registration/registro.html', data) """


import random
from django.core.mail import send_mail
from django.conf import settings

def registro(request):
    if request.method == 'POST':
        user_creation_form = CustomUserCreationForm(
            data=request.POST, files=request.FILES)

        if user_creation_form.is_valid():
            # Generar código aleatorio de 4 dígitos
            verification_code = str(random.randint(1000, 9999))
            # Almacenar datos temporales en la sesión
            request.session['verification_code'] = verification_code
            request.session['user_form_data'] = user_creation_form.cleaned_data
            request.session['user_files'] = request.FILES

            # Enviar email
            email_text = 'Bienvenido al Centro Fray Bartolome de las Casas, para completar su registro ingrese el siguiente codigo : ' + verification_code
            try:
                send_mail(
                    'Código de Verificación - Centro Fray Bartolome de las Casas',
                    email_text,
                    settings.DEFAULT_FROM_EMAIL,
                    [user_creation_form.cleaned_data['email']],
                    fail_silently=False,
                )
                # Redirigir a la página de verificación
                return redirect('principal:verify_email')
            except Exception as e:
                print(f"Error al enviar email: {str(e)}")
                messages.error(request, 'Error al enviar el código de verificación. Por favor, intente nuevamente más tarde.')
        else:
            # Mostrar errores específicos como mensajes
            for field, errors in user_creation_form.errors.items():
                for error in errors:
                    if field == 'password2' and 'password_mismatch' in error:
                        messages.error(request, 'Las contraseñas no coinciden. Por favor, asegúrese de escribir la misma contraseña en ambos campos.')
                    else:
                        messages.error(request, f"{field}: {error}")
            print(f"Errores en el formulario: {user_creation_form.errors}")
    else:
        user_creation_form = CustomUserCreationForm()

    data = {
        'form': user_creation_form
    }
    return render(request, 'registration/registro.html', data)

def verify_email(request):
    if request.method == 'POST':
        code = request.POST.get('code')
        if code == request.session.get('verification_code'):
            user_form_data = request.session.get('user_form_data')
            user_files = request.session.get('user_files')
            user_creation_form = CustomUserCreationForm(data=user_form_data, files=user_files)
            if user_creation_form.is_valid():
                user = user_creation_form.save(commit=True)
                messages.success(request, f"Usuario {user.username} creado correctamente")
                del request.session['verification_code']
                del request.session['user_form_data']
                del request.session['user_files']

                # Enviar correo de confirmación de registro
                confirmation_subject = 'Registro Exitoso - Centro Fray Bartolome de las Casas'
                confirmation_message = f'Usted se ha registrado satisfactoriamente. Su Nombre de Usuario es: {user.username}'
                try:
                    send_mail(
                        confirmation_subject,
                        confirmation_message,
                        settings.DEFAULT_FROM_EMAIL,
                        [user.email],
                        fail_silently=False,
                    )
                except Exception as e:
                    print(f"Error al enviar correo de confirmación: {str(e)}")
                    # Considerar si se debe mostrar un mensaje de error al usuario o simplemente loguear

                return redirect('login')
            else:
                error_message = 'Error al crear el usuario. Por favor, intente nuevamente.'
        else:
            error_message = 'Código incorrecto. Por favor, intente nuevamente.'

        return render(request, 'registration/verify_email.html', {'error': error_message})

    return render(request, 'registration/verify_email.html')

# Vista para manejar la redirección después del login

class LoginRedirectView(LoginRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        user = request.user
        if user.is_authenticated:
            if user.groups.filter(name='Profesores').exists() or user.groups.filter(name='Administración').exists() or user.groups.filter(name='Secretaría').exists():
                return redirect('principal:profile')  # Redirige a la página de perfil del profesor o el admin
            else:
                return redirect('principal:cursos')  # Redirige a la página de cursos para otros usuarios
        return redirect('principal:home') # Redirige a home si no está autenticado (aunque LoginRequiredMixin ya lo manejaría)



    # Pagina de Perfil


class ProfileView(BaseContextMixin, TemplateView):
    template_name = 'profile/profile.html'

    @override
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        if user.groups.first().name == 'Profesores':
            # Obtener todos los cursos asignados al profesor del curso académico activo
            curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
            if curso_academico_activo:
                assigned_courses = Curso.objects.filter(teacher=user, curso_academico=curso_academico_activo)
            else:
                assigned_courses = Curso.objects.none()
            context['assigned_courses'] = assigned_courses
        elif user.groups.first().name == 'Estudiantes':
            # Obtener los cursos en los que el estudiante está inscrito y que pertenecen al curso académico activo
            curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
            if curso_academico_activo:
                enrolled_courses = Curso.objects.filter(matriculas__student=user, curso_academico=curso_academico_activo)
            else:
                enrolled_courses = Curso.objects.none()
            context['enrolled_courses'] = enrolled_courses
        elif user.groups.first().name in ['Administración', 'Secretaría']:
            curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
            if curso_academico_activo:
                all_courses = Curso.objects.filter(curso_academico=curso_academico_activo)
            else:
                all_courses = Curso.objects.none()
            context['all_courses'] = all_courses

        return context

# Vista de los Cursos


class MatriculasListView(BaseContextMixin, ListView):
    model = Matriculas
    template_name = 'matriculas_list.html'
    context_object_name = 'matriculas'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtering by CursoAcademico
        curso_academico_id = self.request.GET.get('curso_academico')
        if curso_academico_id:
            queryset = queryset.filter(curso_academico__id=curso_academico_id)

        # Filtering by Curso
        curso_id = self.request.GET.get('curso')
        if curso_id:
            queryset = queryset.filter(course__id=curso_id)

        # Filtering by Estudiante
        student_id = self.request.GET.get('student')
        if student_id:
            queryset = queryset.filter(student__id=student_id)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cursos_academicos'] = CursoAcademico.objects.all()
        context['cursos'] = Curso.objects.all()
        context['estudiantes'] = User.objects.filter(groups__name='Estudiantes')
        return context


# Vistas para Calificaciones
class CalificacionesListView(BaseContextMixin, ListView):
    model = Calificaciones
    template_name = 'calificaciones_list.html'
    context_object_name = 'calificaciones'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtering by CursoAcademico
        curso_academico_id = self.request.GET.get('curso_academico')
        if curso_academico_id:
            queryset = queryset.filter(curso_academico__id=curso_academico_id)

        # Filtering by Curso
        curso_id = self.request.GET.get('curso')
        if curso_id:
            queryset = queryset.filter(course__id=curso_id)

        # Filtering by Estudiante
        student_id = self.request.GET.get('student')
        if student_id:
            queryset = queryset.filter(student__id=student_id)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cursos_academicos'] = CursoAcademico.objects.all()
        context['cursos'] = Curso.objects.all()
        context['estudiantes'] = User.objects.filter(groups__name='Estudiantes')
        return context


class StudentCourseAttendanceView(BaseContextMixin, ListView):
    model = Asistencia
    template_name = 'student_asistencias.html'
    context_object_name = 'asistencias'

    def get_queryset(self):
        queryset = super().get_queryset()
        student_id = self.kwargs['student_id']
        course_id = self.kwargs['course_id']
        queryset = queryset.filter(student__id=student_id, course__id=course_id)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student_id = self.kwargs['student_id']
        course_id = self.kwargs['course_id']
        context['student'] = User.objects.get(id=student_id)
        context['course'] = Curso.objects.get(id=course_id)
        context['student'] = User.objects.get(id=student_id)
        return context


# Vistas para Asistencias
class AsistenciasListView(ListView):
    model = Asistencia
    template_name = 'asistencias_list.html'
    context_object_name = 'asistencias'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtrar por Curso Académico
        curso_academico_id = self.request.GET.get('curso_academico')
        if curso_academico_id:
            queryset = queryset.filter(course__curso_academico__id=curso_academico_id)

        # Filtrar por Curso
        curso_id = self.request.GET.get('curso')
        if curso_id:
            queryset = queryset.filter(course__id=curso_id)

        # Filtrar por Estudiante
        estudiante_id = self.request.GET.get('estudiante')
        if estudiante_id:
            queryset = queryset.filter(student__id=estudiante_id)

        # Filtrar por Fecha
        fecha_asistencia = self.request.GET.get('fecha')
        if fecha_asistencia:
            queryset = queryset.filter(date=fecha_asistencia)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cursos_academicos'] = CursoAcademico.objects.all()
        context['cursos'] = Curso.objects.all()
        context['estudiantes'] = User.objects.filter(groups__name='Estudiantes')
        context['selected_curso_academico'] = self.request.GET.get('curso_academico')
        
        # Calcular porcentaje de asistencia cuando se filtra por curso y estudiante
        curso_id = self.request.GET.get('curso')
        estudiante_id = self.request.GET.get('estudiante')
        
        if curso_id and estudiante_id:
            # Obtener todas las asistencias del estudiante en el curso
            total_asistencias = Asistencia.objects.filter(course_id=curso_id, student_id=estudiante_id).count()
            presentes = Asistencia.objects.filter(course_id=curso_id, student_id=estudiante_id, presente=True).count()
            
            if total_asistencias > 0:
                porcentaje = (presentes / total_asistencias) * 100
                context['porcentaje_asistencia'] = round(porcentaje, 2)
                context['total_asistencias'] = total_asistencias
                context['presentes'] = presentes
                context['ausentes'] = total_asistencias - presentes
        context['selected_curso'] = self.request.GET.get('curso')
        context['selected_estudiante'] = self.request.GET.get('estudiante')
        context['selected_fecha'] = self.request.GET.get('fecha')
        return context


class StudentCourseNotesView(BaseContextMixin, ListView):
    model = Calificaciones
    template_name = 'student_notes.html'  # New template for student notes
    context_object_name = 'calificaciones'

    def get_queryset(self):
        student_id = self.kwargs['student_id']
        course_id = self.kwargs['course_id']
        # Filter grades for the specific student and course
        return Calificaciones.objects.filter(student__id=student_id, course__id=course_id)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student_id = self.kwargs['student_id']
        course_id = self.kwargs['course_id']

        student = User.objects.get(id=student_id)
        course = Curso.objects.get(id=course_id)

        context['student'] = student
        context['course'] = course

        print(f"[DEBUG] StudentCourseNotesView - student_id: {student_id}, course_id: {course_id}")

        # Get all Calificaciones for the student and course
        calificaciones_for_student_course = Calificaciones.objects.filter(
            student=student,
            course=course
        )
        print(f"[DEBUG] Calificaciones found: {calificaciones_for_student_course.count()}")

        all_notes = []
        total_score = 0
        num_grades = 0

        for calificacion in calificaciones_for_student_course:
            print(f"[DEBUG] Processing Calificacion ID: {calificacion.id}")
            for nota in calificacion.notas.all():
                print(f"[DEBUG] Adding Nota: {nota.valor} (ID: {nota.id})")
                all_notes.append(nota)
                if nota.valor is not None:
                    total_score += nota.valor
                    num_grades += 1

        if num_grades > 0:
            average_score = total_score / num_grades
        else:
            average_score = 0

        context['all_notes'] = all_notes
        context['average_score'] = average_score
        return context


class CoursesView(BaseContextMixin, TemplateView):
    template_name = 'cursos.html'

    @override
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
        if curso_academico_activo:
            courses = Curso.objects.filter(curso_academico=curso_academico_activo)
        else:
            courses = Curso.objects.none()
        student = self.request.user if self.request.user.is_authenticated else None

        for item in courses:
            if student:
                registration = Matriculas.objects.filter(
                    course=item, student=student).first()
                item.is_enrolled = registration is not None
            else:
                item.is_enrolled = False

            # Calcular el conteo de inscripciones para todos los cursos
            enrollment_count = Matriculas.objects.filter(course=item).count()
            item.enrollment_count = enrollment_count

        context['courses'] = courses
        # Asegurarse de que group_name esté en el contexto
        user = self.request.user
        if user.is_authenticated:
            group = Group.objects.filter(user=user).first()
            if group:
                context['group_name'] = group.name
            else:
                context['group_name'] = None
        else:
            context['group_name'] = None

        return context

# Crear nuevo Curso


class CourseCreateView(LoginRequiredMixin, CreateView):
    model = Curso
    form_class = CourseForm
    template_name = 'create_course.html'
    success_url = reverse_lazy('principal:cursos')

    def form_valid(self, form):
        # Asigna el profesor actual al curso antes de guardar
        form.instance.teacher = self.request.user
        # Asigna el curso académico activo al curso
        active_academic_course = CursoAcademico.objects.filter(activo=True).first()
        if active_academic_course:
            form.instance.curso_academico = active_academic_course
        messages.success(self.request, 'El Curso se guardo correctamente')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(
            self.request, 'Ha ocurrido un ERROR al guardar el Curso')
        return self.render_to_response(self.get_context_data(form=form))

# Vista para inscribirse a un curso


@login_required
def inscribirse_curso(request, curso_id):
    # Obtener el curso
    curso = Curso.objects.get(id=curso_id)
    estudiante = request.user

    # Verificar si ya está inscrito
    inscripcion_existente = Matriculas.objects.filter(
        course=curso, student=estudiante).exists()

    if not inscripcion_existente:
        # Obtener el curso académico activo
        curso_academico = CursoAcademico.objects.filter(activo=True).first()
        
        if not curso_academico:
            messages.error(request, 'No hay un curso académico activo configurado. Contacte al administrador.')
            return redirect('principal:cursos')
            
        # Crear nueva matrícula asignada al curso académico activo
        matricula = Matriculas(
            course=curso, 
            student=estudiante, 
            activo=True,
            curso_academico=curso_academico,
            estado='P'  # Estado inicial: Pendiente
        )
        matricula.save()
        messages.success(
            request, f'Te has inscrito exitosamente al curso {curso.name} para el año académico {curso_academico.nombre}')
    else:
        messages.info(request, 'Ya estás inscrito en este curso')

    return redirect('principal:cursos')

# Vista para editar un curso


class CourseUpdateView(BaseContextMixin, UpdateView):
    model = Curso
    form_class = CourseForm
    template_name = 'create_course.html'  # Reutilizamos el mismo template
    success_url = reverse_lazy('principal:cursos')

    @override
    def form_valid(self, form):
        messages.success(self.request, 'El Curso se actualizó correctamente')
        return super().form_valid(form)

    @override
    def form_invalid(self, form):
        messages.error(
            self.request, 'Ha ocurrido un ERROR al actualizar el Curso')
        return self.render_to_response(self.get_context_data(form=form))


# Vista para eliminar un curso
@login_required
def eliminar_curso(request, curso_id):
    # Verificar si el usuario pertenece al grupo 'Secretaría'
    if request.user.groups.filter(name='Secretaría').exists():
        try:
            # Obtener el curso
            curso = Curso.objects.get(id=curso_id)
            nombre_curso = curso.name

            # Eliminar el curso
            curso.delete()
            messages.success(
                request, f'El curso {nombre_curso} ha sido eliminado correctamente')
        except Curso.DoesNotExist:
            messages.error(request, 'El curso no existe')
    else:
        messages.error(request, 'No tienes permisos para eliminar cursos')

    return redirect('principal:cursos')

# Mostrar lista de alumnos y notas a los profesores


class StudentListNotasView(BaseContextMixin, ListView):
    model = Matriculas
    template_name = 'student_list_notas.html'
    context_object_name = 'matriculas'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'student',
            'course',
            'course__teacher',
            'calificaciones'
        )

        # Verificar si se está accediendo desde la URL con course_id
        course_id = self.kwargs.get('course_id')
        if course_id:
            queryset = queryset.filter(course__id=course_id)
            return queryset

        # Si no hay course_id en la URL, usar los filtros normales
        search_query = self.request.GET.get('search_query')
        course_filter = self.request.GET.get('course')
        teacher_filter = self.request.GET.get('teacher')

        if search_query:
            queryset = queryset.filter(
                Q(student__username__icontains=search_query) |
                Q(student__first_name__icontains=search_query) |
                Q(student__last_name__icontains=search_query)
            )


        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        course_id = self.kwargs.get('course_id')
        if course_id:
            course = get_object_or_404(Curso, id=course_id)
            context['course'] = course

            # Obtener el curso académico activo
            curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
            context['curso_academico'] = curso_academico_activo

            # Obtener solo las matrículas activas para este curso y curso académico activo
            active_enrollments = Matriculas.objects.filter(
                course=course,
                activo=True,
                curso_academico=curso_academico_activo
            )

            student_data = []
            for enrollment in active_enrollments:
                student = enrollment.student
                # Buscar calificación por curso, estudiante y curso académico
                calificacion = Calificaciones.objects.filter(
                    course=course,
                    student=student,
                    curso_academico=curso_academico_activo
                ).first()

                all_notes = []
                if calificacion:
                    # Obtener todas las notas individuales relacionadas con esta calificación
                    all_notes = list(calificacion.notas.all().order_by('fecha_creacion'))

                student_data.append({
                    'calificacion_id': calificacion.id if calificacion else None,
                    'name': student.get_full_name(),
                    'notas': all_notes, # Lista de notas individuales
                    'average': calificacion.average if calificacion else None,
                    'matricula_id': enrollment.id,
                    'student_id': student.id, # Add student.id here
                })
            context['student_data'] = student_data
        else:
            context['courses'] = CursoAcademico.objects.all()
            context['teachers'] = User.objects.filter(groups__name='Docente')
        
        # Obtener el curso académico activo
        curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
        
        # Obtener solo las matrículas activas para este curso y curso académico activo
        active_enrollments = Matriculas.objects.filter(
            course=course, 
            activo=True,
            curso_academico=curso_academico_activo
        )
        
        student_data = []
        for enrollment in active_enrollments:
            student = enrollment.student
            # Buscar calificación por curso, estudiante y curso académico
            calificacion = Calificaciones.objects.filter(
                course=course, 
                student=student,
                curso_academico=curso_academico_activo
            ).first()
            
            all_notes = []
            notas_list = []
            if calificacion:
                # Obtener todas las notas individuales relacionadas con esta calificación
                notas_list = list(calificacion.notas.all().order_by('fecha_creacion').values_list('valor', flat=True))

            student_data.append({
                'calificacion_id': calificacion.id if calificacion else None,
                'name': student.get_full_name(),
                'notas': notas_list, # Lista de notas individuales
                'average': calificacion.average if calificacion else None,
                'matricula_id': enrollment.id,
                'student_id': student.id, # Add student.id here
            })

        context['course'] = course
        context['student_data'] = student_data
        context['curso_academico'] = curso_academico_activo
        return context
# Agregar Notas de los estudiantes

class AddNotaView(LoginRequiredMixin, View):
    def dispatch(self, request, *args, **kwargs):

        return super().dispatch(request, *args, **kwargs)

    def get(self, request, matricula_id):
        matricula = get_object_or_404(Matriculas, id=matricula_id)
        print(f"[GET] Matricula ID: {matricula_id}")
        print(f"[GET] Matricula Course ID: {matricula.course.id}")
        print(f"[GET] Matricula Student ID: {matricula.student.id}")
        print(f"[GET] Matricula Curso Academico ID: {matricula.curso_academico.id if matricula.curso_academico else 'None'}")

        try:
            # Buscar calificación por curso, estudiante y curso académico de la matrícula
            calificacion = Calificaciones.objects.get(
                course=matricula.course, 
                student=matricula.student,
                curso_academico=matricula.curso_academico
            )
            print("[GET] Existing Calificacion found.")
            form = CalificacionesForm(instance=calificacion)
            formset = NotaIndividualFormSet(instance=calificacion) # Instancia el formset con la calificación existente
        except Calificaciones.DoesNotExist:
            print("[GET] Calificaciones.DoesNotExist raised. No existing Calificacion found.")
            form = CalificacionesForm(initial={
                'matricula': matricula,
                'course': matricula.course,
                'student': matricula.student,
                'curso_academico': matricula.curso_academico
            }) # Pasa los valores iniciales para los campos de Calificaciones
            formset = NotaIndividualFormSet() # Instancia un formset vacío
        
        context = {
            'form': form,
            'formset': formset, # Añade el formset al contexto
            'matricula': matricula
        }
        return render(request, 'add_nota.html', context)

    def post(self, request, matricula_id):
        matricula = get_object_or_404(Matriculas, id=matricula_id)
        print(f"[POST] Matricula ID: {matricula_id}")
        print(f"[POST] Matricula Course ID: {matricula.course.id}")
        print(f"[POST] Matricula Student ID: {matricula.student.id}")
        print(f"[POST] Matricula Curso Academico ID: {matricula.curso_academico.id if matricula.curso_academico else 'None'}")

        try:
            calificacion = Calificaciones.objects.get(
                course=matricula.course,
                student=matricula.student,
                curso_academico=matricula.curso_academico
            )
            print("[POST] Existing Calificacion found.")
            form = CalificacionesForm(request.POST, instance=calificacion)
        except Calificaciones.DoesNotExist:
            print("[POST] Calificaciones.DoesNotExist raised. No existing Calificacion found. Creating new one.")
            form = CalificacionesForm(request.POST)

        if form.is_valid():
            calificacion = form.save(commit=False)
            calificacion.matricula = matricula
            calificacion.course = matricula.course
            calificacion.student = matricula.student
            calificacion.curso_academico = matricula.curso_academico
            calificacion.save() # <-- Guarda la instancia de Calificaciones aquí

            # Ahora que calificacion tiene un PK, podemos inicializar el formset
            formset = NotaIndividualFormSet(request.POST, instance=calificacion)

            if formset.is_valid():
                
                formset.save()
                messages.success(request, 'Notas guardadas correctamente.')
                return redirect('principal:student_list_notas_by_course', course_id=matricula.course.id)
            else:
                print(f"[POST] Formset is NOT valid. Errors: {formset.errors}")
                messages.error(request, 'Error al guardar las notas individuales.')
        else:
            print(f"[POST] CalificacionesForm errors: {form.errors}")
            messages.error(request, 'Error al guardar la calificación principal.')

        context = {
            'form': form,
            'formset': formset, # Asegúrate de pasar el formset al contexto incluso si hay errores
            'matricula': matricula
        }
        return render(request, 'add_nota.html', context)

#esto es de la ia
""" def crear_cursos(request):
    if request.method == 'POST':
        form = CourseForm(request.POST, request.FILES) # Asegúrate de incluir request.FILES aquí
        if form.is_valid():
            form.save()
            # Por ejemplo, puedes añadir un mensaje de éxito y redirigir
            # messages.success(request, 'Curso creado exitosamente!')
            # return redirect('nombre_de_tu_url_de_cursos')
    else:
        form = CourseForm()
    return render(request, 'create_course.html', {'form': form})


def editar_curso(request, course_id):
    course = get_object_or_404(Curso, id=course_id)
    if request.method == 'POST':
        form = CourseForm(request.POST, request.FILES, instance=course) # Asegúrate de incluir request.FILES aquí
        if form.is_valid():
            form.save()
            # Por ejemplo, puedes añadir un mensaje de éxito y redirigir
            # messages.success(request, 'Curso actualizado exitosamente!')
            # return redirect('nombre_de_tu_url_de_cursos')
    else:
        form = CourseForm(instance=course)
    return render(request, 'edit_course.html', {'form': form, 'course': course})

 """

#vistas para historicos

def historico_alumno(request, student_id):
        matriculas = Matriculas.objects.filter(student_id=student_id).select_related('curso_academico')
        return render(request, 'historico.html', {'matriculas': matriculas})


# Agregando asistencias

class AsistenciaView(BaseContextMixin, TemplateView):
    template_name = 'asistencias.html'
    
    @override
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        course_id = kwargs['course_id']
        course = get_object_or_404(Curso, id=course_id)
        
        # Obtener el curso académico activo
        curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
        
        # Filtrar asistencias por curso y ordenar por fecha descendente
        asistencias = Asistencia.objects.filter(
            course=course,
            student__matriculas__curso_academico=curso_academico_activo,
            student__matriculas__activo=True
        ).select_related('student', 'course').order_by('-date')
        
        # Obtener todas las matrículas activas para este curso en el curso académico activo
        matriculas = Matriculas.objects.filter(
            course=course,
            activo=True,
            curso_academico=curso_academico_activo
        ).select_related('student')  # Optimizar consulta de estudiantes
        
        # Filtrar por fecha si se proporciona en la solicitud
        fecha_filtro = self.request.GET.get('fecha')
        if fecha_filtro:
            asistencias = asistencias.filter(date=fecha_filtro)
        
        # Calcular la cantidad de asistencias registradas (fechas únicas)
        asistencias_registradas = Asistencia.objects.filter(course=course).values('date').distinct().count()
        
        # Obtener la cantidad total de clases del curso
        cantidad_total_clases = course.class_quantity
        
        # Calcular las clases restantes
        clases_restantes = cantidad_total_clases - asistencias_registradas

        context['course'] = course
        context['asistencias'] = asistencias
        context['matriculas'] = matriculas
        context['curso_academico'] = curso_academico_activo
        context['cantidad_total_clases'] = cantidad_total_clases
        context['asistencias_registradas'] = asistencias_registradas
        context['clases_restantes'] = clases_restantes
        return context


class AddAsistenciaView(LoginRequiredMixin, TemplateView):
    template_name='add_asistencias.html'

    @override
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        course_id = kwargs['course_id']
        course = Curso.objects.get(id=course_id)
        matriculas = Matriculas.objects.filter(course=course)
        # Obtener las asistencias existentes para este curso
        asistencias = Asistencia.objects.order_by('date')
        context['course'] = course
        context['matriculas'] = matriculas
        context['asistencias'] = asistencias
        context['today'] = date.today()
        return context

    def post(self, request, course_id):
        course = Curso.objects.get(id=course_id)
        matriculas = Matriculas.objects.filter(course=course)

        if request.method == 'POST':
            date = request.POST.get('date')

            for matricula in matriculas:
                absent = request.POST.get('asistencia_' + str(matricula.id))
                # Buscar si ya existe un registro de asistencia para este estudiante en esta fecha
                asistencia, created = Asistencia.objects.get_or_create(
                    student=matricula.student,
                    course=course,
                    date=date,
                    defaults={'presente': not bool(absent)}
                )
                # Si el registro ya existía, actualizar el estado de presente
                if not created:
                    asistencia.presente = not bool(absent)
                    asistencia.save()
        # Redirigir a la misma página para mostrar las asistencias actualizadas
        return redirect('principal:asistencias', course_id=course_id)


# Eliminar asistencia
def eliminar_asistencia(request, asistencia_id):
    # Obtener la asistencia o devolver 404 si no existe
    asistencia = get_object_or_404(Asistencia, id=asistencia_id)
    
    # Guardar el ID del curso antes de eliminar la asistencia
    course_id = asistencia.course.id
    
    # Eliminar la asistencia
    asistencia.delete()
    
    # Redirigir a la página de asistencias del curso
    return redirect('principal:asistencias', course_id=course_id)


def add_asistencias(request, course_id):
    course = get_object_or_404(Curso, id=course_id)
    matriculas = Matriculas.objects.filter(course=course, activo=True)

    if request.method == 'POST':
        date_str = request.POST.get('date')
        try:
            attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Formato de fecha inválido.")
            return redirect('principal:add_asistencias', course_id=course.id)

        # Eliminar asistencias existentes para esta fecha y curso
        Asistencia.objects.filter(course=course, date=attendance_date).delete()

        for matricula in matriculas:
            is_absent = request.POST.get(f'asistencia_{matricula.id}')
            presente = not bool(is_absent) # Si está marcado, significa que está ausente, por lo tanto, no presente

            Asistencia.objects.create(
                course=course,
                student=matricula.student,
                date=attendance_date,
                presente=presente
            )
        messages.success(request, "Asistencias guardadas correctamente.")
        return redirect('principal:asistencias', course_id=course.id) # Redirige a la página de asistencias del curso
    
    today = date.today()
    context = {
        'course': course,
        'matriculas': matriculas,
        'today': today,
    }
    return render(request, 'add_asistencias.html', context)


@login_required
def undo_last_asistencia(request, course_id):
    course = get_object_or_404(Curso, id=course_id)

    # Encontrar la fecha más reciente para la que se registró asistencia en este curso
    latest_attendance_date_obj = Asistencia.objects.filter(course=course).aggregate(Max('date'))
    latest_attendance_date = latest_attendance_date_obj['date__max']

    if latest_attendance_date:
        # Eliminar todos los registros de asistencia para este curso en la fecha más reciente
        Asistencia.objects.filter(course=course, date=latest_attendance_date).delete()
        messages.success(request, f"La asistencia del {latest_attendance_date.strftime('%d-%m-%Y')} ha sido deshecha correctamente.")
    else:
        messages.info(request, "No hay asistencias registradas para deshacer en este curso.")

    return redirect('principal:asistencias', course_id=course.id)

    

